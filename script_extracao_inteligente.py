import openpyxl
import re
from tqdm import tqdm

arquivo = '01 - Janeiro.xlsx'
wb = openpyxl.load_workbook(arquivo)
ws = wb.active

ultima_coluna = ws.max_column

# Definir novos cabeçalhos a partir da linha 2
cabecalhos = ['Padrão', 'Código', 'Empresa/Pessoa', 'Descrição', 'Pedido', 'Observação']
for idx, cabecalho in enumerate(cabecalhos, start=1):
    ws.cell(row=2, column=ultima_coluna + idx).value = cabecalho

# Função inteligente de identificação dinâmica de padrões
def identificar_padroes(texto):
    texto_original = texto  # manter texto original para observações
    padrao, codigo, empresa, descricao, pedido, observacao = '', '', '', '', '', ''

    # 01. Nota fiscal (NF.)
    if texto.startswith('NF.'):
        padrao = 'Nota Fiscal'
        partes = texto.split(' - ')
        if len(partes) >= 4:
            codigo_empresa = partes[0].split('/')
            codigo = codigo_empresa[0].strip()
            empresa = codigo_empresa[1].strip() if len(codigo_empresa) > 1 else ''
            descricao = partes[1].strip()
            pedido = partes[2].strip()
            observacao = ' - '.join(partes[3:]).strip()

    # 02. MOV BANC PAG
    elif texto.startswith('MOV BANC PAG.'):
        padrao = 'Movimentação Bancária Pagamento'
        partes = texto.split(' - ')
        codigo = partes[0].strip()
        descricao = partes[1].strip() if len(partes) > 1 else ''
        observacao = ' - '.join(partes[2:]).strip() if len(partes) > 2 else ''

    # 03. Apropriação seguro veículos ou máquinas
    elif 'APROPRIACAO DE SEGURO VEICULOS' in texto or 'SEGURO MAQUINAS E EQUIPAMENTOS' in texto:
        padrao = 'Apropriação Seguro'
        partes = texto.split(' - ')
        descricao = partes[0].strip()
        empresa = partes[1].strip() if len(partes) > 1 else ''

    # 04. Estornos
    elif texto.startswith('ESTORNO'):
        padrao = 'Estorno'
        descricao = texto.strip()

    # 05. Apropriações específicas (IPVA, Licenciamento)
    elif 'APROPRIACAO IPVA' in texto or 'APROPRIACAO LICENCIAMENTO' in texto:
        padrao = 'Apropriação IPVA/Licenciamento'
        descricao = texto.strip()

    # 06. Requisições específicas
    elif texto.startswith('REQ. TIPO:'):
        padrao = 'Requisição'
        partes = texto.split(' - ')
        descricao = partes[0].strip()
        observacao = partes[1].strip() if len(partes) > 1 else ''

    # 07. Transferências específicas (TRF)
    elif texto.startswith('TRF'):
        padrao = 'Transferência'
        descricao = texto.strip()

    # 08. Códigos internos folha pagamento (Pxxx ou Dxxx)
    elif re.match(r'^[PD]\d{3}', texto):
        padrao = 'Folha Pagamento Interna'
        partes = texto.split('-')
        codigo = partes[0].strip()
        descricao = partes[1].strip() if len(partes) > 1 else ''

    # 09. Códigos folha pagamento (4 dígitos numéricos)
    elif re.match(r'^\d{4}-', texto):
        padrao = 'Folha Pagamento Numérica'
        partes = texto.split('-')
        codigo = partes[0].strip()
        descricao = partes[1].strip() if len(partes) > 1 else ''

    # 10. Códigos folha numérica especial (ex: 0100-13º)
    elif re.match(r'^\d{4}-', texto):
        padrao = 'Folha Numérica Especial'
        partes = texto.split('-')
        codigo = partes[0].strip()
        descricao = partes[1].strip() if len(partes) > 1 else ''

    # 11. BEM (Patrimônio)
    elif texto.startswith('BEM'):
        padrao = 'Patrimônio'
        partes = texto.split(' - ')
        codigo = partes[0].strip()
        descricao = partes[1].strip() if len(partes) > 1 else ''

    # 12. Demais não classificados
    else:
        padrao = 'Outros'
        descricao = texto.strip()

    return padrao, codigo, empresa, descricao, pedido, observacao

# Executar extração com barra de progresso
for linha in tqdm(range(3, ws.max_row + 1), desc="Processando linhas"):
    texto_celula = ws.cell(row=linha, column=9).value
    if texto_celula:
        resultado = identificar_padroes(texto_celula)
        for idx, valor in enumerate(resultado, start=1):
            ws.cell(row=linha, column=ultima_coluna + idx).value = valor

# Salvar arquivo com extração inteligente
wb.save('01 - Janeiro_Extraido_Inteligente.xlsx')

print("✅ Extração inteligente concluída com sucesso!")



